# -*- coding: utf-8 -*-
# Membangun Excel komparasi waktu RQ-A (durasi transaksi sebelum vs sesudah POS).
# Gaya tabel mengikuti contoh thesis (Metode|Skenario|Waktu + ringkasan efisiensi).
# Angka konsisten dgn model di rqa-simulasi-durasi.mjs (parameter sama).
# Jalankan: python docs/pengujian/build-rqa-excel.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r'docs/pengujian/RQ-A-komparasi-waktu.xlsx'

# ---- gaya ----
HDR = PatternFill('solid', fgColor='1F6F43')      # hijau tua
HDRF = Font(bold=True, color='FFFFFF', size=11)
TITLE = Font(bold=True, size=13, color='1F6F43')
SUB = Font(bold=True, size=11)
NOTE = Font(italic=True, size=9, color='555555')
MANUALF = PatternFill('solid', fgColor='FBE9E7')  # merah muda
SISF = PatternFill('solid', fgColor='E8F5E9')     # hijau muda
WIN = Font(bold=True, color='1B5E20')             # hijau (POS cepat)
LOSE = Font(bold=True, color='B71C1C')            # merah (POS lambat)
thin = Side(style='thin', color='BBBBBB')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')
CTR = Alignment(horizontal='center', vertical='center')

def style_header(ws, row, c1, c2):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = CTR; cell.border = BORD

def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = BORD

wb = openpyxl.Workbook()

# =========================================================================
# SHEET 1 — Komparasi Aktivitas (gaya foto: Metode | Skenario | Waktu)
# =========================================================================
ws = wb.active; ws.title = 'Komparasi Aktivitas'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 62
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 34

r = 1
ws.cell(r,1,'Komparasi Waktu Proses Transaksi — Sebelum (Manual) vs Sesudah (POS)').font = TITLE; r+=2
intro = ('Pesanan acuan untuk Tabel 1–3: 1 transaksi dine-in berisi 4 jenis item + 1 paket, bayar tunai '
         '(mewakili pesanan kompleks yang lazim di resto). Waktu = HASIL MODEL (penjumlahan waktu per-langkah '
         '× komposisi nyata), bukan stopwatch; hanya latensi yang diukur nyata ke server. "Sistem (terbiasa)" '
         '= pegawai sudah mahir; angka masa adaptasi (awam) ada di sheet "Ringkasan Efisiensi". '
         'Detail parameter & asal-angka di sheet terpisah.')
ws.cell(r,1,intro); ws.cell(r,1).font = NOTE; ws.cell(r,1).alignment = WRAP
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); ws.row_dimensions[r].height=58; r+=2

def activity_table(r, no, judul, rows, efis=None):
    ws.cell(r,1, f'Tabel {no} Komparasi Waktu {judul}').font = SUB; r+=1
    for i,h in enumerate(['Metode','Skenario','Waktu (detik)','Catatan'],start=1):
        ws.cell(r,i,h)
    style_header(ws, r, 1, 4); r+=1
    start=r
    for metode, skenario, waktu, catatan, fill, ftan in rows:
        ws.cell(r,1,metode).alignment=WRAP
        ws.cell(r,2,skenario).alignment=WRAP
        wc=ws.cell(r,3,waktu); wc.alignment=CTR
        ws.cell(r,4,catatan).alignment=WRAP; ws.cell(r,4).font=NOTE
        if fill:
            for c in range(1,5): ws.cell(r,c).fill=fill
        if ftan: ws.cell(r,3).font=ftan
        ws.row_dimensions[r].height=42
        r+=1
    box(ws, start, 1, r-1, 4)
    if efis is not None:
        ws.cell(r,1,'Efisiensi waktu'); ws.cell(r,1).font=SUB
        ws.cell(r,2,f'= (Manual − Sistem terbiasa) / Manual × 100%')
        ce=ws.cell(r,3,efis); ce.font=(LOSE if str(efis).startswith('−') else WIN); ce.alignment=CTR
        box(ws, r,1,r,4); r+=1
    return r+1

r = activity_table(r, '4.x.1', 'Pencatatan Pesanan oleh Pelayan', [
    ('Manual','Pelayan menulis pesanan (4 item + 1 paket) di kertas, lalu menyerahkan ke kasir.','27', '', MANUALF, None),
    ('Sistem (terbiasa)','Pelayan membuka meja, men-tap 4 item + memilih opsi paket, lalu klik Simpan.','36', 'POS LEBIH LAMBAT (+9 dtk): input lewat layar > tulis tangan.', SISF, LOSE),
])
r = activity_table(r, '4.x.2', 'Pembuatan Nota & Pembayaran oleh Kasir', [
    ('Manual','Kasir menyalin pesanan ke nota, menjumlahkan harga manual, menerima tunai & menghitung kembalian.','52', '', MANUALF, None),
    ('Sistem (terbiasa)','Kasir membuka meja, modal checkout (item & total otomatis), input tunai (kembalian otomatis), cetak struk.','10,6', 'POS JAUH LEBIH CEPAT (−41 dtk): hitung & kembalian otomatis.', SISF, WIN),
], efis='80%')
r = activity_table(r, '4.x.3', 'Total per Transaksi (Pelayan + Kasir)', [
    ('Manual','Tulis kertas → serah → kasir salin+hitung → bayar tunai.','79', '', MANUALF, None),
    ('Sistem (terbiasa)','Input POS → bayar (total & kembalian otomatis) → struk.','46,6', 'Beban bergeser: pelayan +, kasir −−; NET lebih cepat.', SISF, WIN),
], efis='41%')
r = activity_table(r, '4.x.4', 'Pembayaran Gabungan (Split-tender, 1 meja 2 metode)', [
    ('Manual','Kasir buka kalkulator, hitung sisa (total − bayar pertama), catat 2 baris pembayaran.','93', '', MANUALF, None),
    ('Sistem (terbiasa)','Toggle "Bayar Sebagian", input tunai → sisa tampil otomatis → lanjut QRIS sisa tagihan.','47,8', 'Sisa dihitung otomatis (tanpa kalkulator).', SISF, WIN),
], efis='48,6%')
r = activity_table(r, '4.x.5', 'Penggabungan Meja (Merge 2 meja)', [
    ('Manual','Kasir jumlahkan total meja-1 + meja-2 (kalkulator), tandai pembayaran gabung.','76', '', MANUALF, None),
    ('Sistem (terbiasa)','Kasir pencet Merge, pilih meja, total gabungan tampil otomatis.','60,8', 'Hemat tipis (hanya 1 penjumlahan antar-meja).', SISF, WIN),
], efis='19,9%')
r = activity_table(r, '4.x.6', 'Gabung Meja + Split-tender (skenario terberat manual)', [
    ('Manual','Jumlah meja-1+meja-2 (kalkulator), tandai gabung, hitung sisa split (kalkulator), catat 2 baris.','110', '', MANUALF, None),
    ('Sistem (terbiasa)','Merge → total otomatis → Bayar Sebagian → sisa otomatis → struk.','66', 'Dua beban kalkulator manual dihapus.', SISF, WIN),
], efis='40%')

# =========================================================================
# SHEET 2 — Durasi 28 Transaksi
# =========================================================================
ws2 = wb.create_sheet('Durasi 28 Transaksi')
for col,w in zip('ABCDEFGHI',[8,6,12,8,7,10,12,12,12]):
    ws2.column_dimensions[col].width = w
ws2.cell(1,1,'Durasi per Transaksi (28 transaksi riil 21–27 Mei 2026, jaringan 4G)').font = TITLE
ws2.cell(2,1,'Δ = Manual − POS (positif = POS lebih cepat; negatif = POS lebih lambat). Kategori: Sederhana = ≤2 jenis item & tanpa paket.')
ws2.cell(2,1).font = NOTE
hdr = ['Tgl','No','Kategori','Jenis item','Paket','Metode','Manual (s)','POS terbiasa (Δ)','POS awam (Δ)']
for i,h in enumerate(hdr,start=1): ws2.cell(4,i,h)
style_header(ws2,4,1,9)
# data: (tgl,no,kat,lines,paket,metode,manual,posM,dm,posN,dn)
DATA = [
 ('21',1,'Sederhana',1,0,'transfer',30.0,26.6,3.4,32.6,-2.6),('21',2,'Kompleks',4,0,'qris',68.0,42.6,25.4,57.6,10.4),
 ('21',3,'Sederhana',1,0,'cash',36.0,22.6,13.4,28.6,7.4),('21',4,'Kompleks',4,3,'cash',85.0,58.6,26.4,82.6,2.4),
 ('21',5,'Kompleks',4,0,'qris',68.0,42.6,25.4,57.6,10.4),('22',1,'Kompleks',2,1,'qris',47.0,36.6,10.4,48.6,-1.6),
 ('22',2,'Kompleks',2,2,'qris',50.0,42.6,7.4,57.6,-7.6),('22',3,'Sederhana',1,0,'edc',32.0,28.6,3.4,34.6,-2.6),
 ('22',4,'Kompleks',6,0,'qris',92.0,54.6,37.4,75.6,16.4),('23',1,'Sederhana',1,0,'qris',28.0,24.6,3.4,30.6,-2.6),
 ('23',2,'Kompleks',4,1,'qris',71.0,48.6,22.4,66.6,4.4),('23',3,'Kompleks',3,2,'qris',62.0,48.6,13.4,66.6,-4.6),
 ('23',4,'Kompleks',8,0,'qris',116.0,66.6,49.4,93.6,22.4),('23',5,'Kompleks',4,1,'qris',71.0,48.6,22.4,66.6,4.4),
 ('23',6,'Kompleks',4,0,'qris',68.0,42.6,25.4,57.6,10.4),('23',7,'Kompleks',10,0,'qris',140.0,78.6,61.4,111.6,28.4),
 ('24',1,'Kompleks',6,0,'cash',100.0,52.6,47.4,73.6,26.4),('24',2,'Sederhana',1,0,'edc',32.0,28.6,3.4,34.6,-2.6),
 ('24',3,'Kompleks',1,1,'qris',31.0,30.6,0.4,39.6,-8.6),('25',1,'Sederhana',1,0,'qris',28.0,24.6,3.4,30.6,-2.6),
 ('25',2,'Kompleks',6,0,'qris',92.0,54.6,37.4,75.6,16.4),('25',3,'Kompleks',1,1,'edc',35.0,34.6,0.4,43.6,-8.6),
 ('25',4,'Kompleks',3,0,'qris',56.0,36.6,19.4,48.6,7.4),('26',1,'Kompleks',2,1,'qris',47.0,36.6,10.4,48.6,-1.6),
 ('26',2,'Sederhana',2,0,'qris',44.0,30.6,13.4,39.6,4.4),('26',3,'Kompleks',1,1,'transfer',33.0,32.6,0.4,41.6,-8.6),
 ('26',4,'Sederhana',1,0,'cash',36.0,22.6,13.4,28.6,7.4),('27',1,'Sederhana',2,0,'transfer',46.0,32.6,13.4,41.6,4.4),
]
rr=5
def fnum(x): return str(x).replace('.',',')
for d in DATA:
    tgl,no,kat,ln,pk,met,man,pm,dm,pn,dn = d
    ws2.cell(rr,1,tgl); ws2.cell(rr,2,no); ws2.cell(rr,3,kat); ws2.cell(rr,4,ln); ws2.cell(rr,5,pk); ws2.cell(rr,6,met)
    ws2.cell(rr,7,fnum(man)).alignment=CTR
    cM=ws2.cell(rr,8,f'{fnum(pm)}  ({"+" if dm>=0 else ""}{fnum(dm)})'); cM.alignment=CTR; cM.font = WIN if dm>1 else (LOSE if dm<-1 else Font())
    cN=ws2.cell(rr,9,f'{fnum(pn)}  ({"+" if dn>=0 else ""}{fnum(dn)})'); cN.alignment=CTR; cN.font = WIN if dn>1 else (LOSE if dn<-1 else Font())
    rr+=1
box(ws2,4,1,rr-1,9)

# =========================================================================
# SHEET 3 — Ringkasan Efisiensi (gaya Toko X Tabel 4.4)
# =========================================================================
ws3 = wb.create_sheet('Ringkasan Efisiensi')
for col,w in zip('ABCDEF',[30,16,16,16,16,16]): ws3.column_dimensions[col].width=w
ws3.cell(1,1,'Ringkasan Efisiensi Waktu Transaksi (rata-rata, detik)').font=TITLE
ws3.cell(2,1,'Efisiensi = (Manual − Sistem) / Manual × 100%. Dua tingkat kemahiran: "terbiasa" (mahir) & "awam" (masa adaptasi, UI ×1,5).')
ws3.cell(2,1).font=NOTE
hd=['Kelompok','Manual (s)','POS terbiasa (s)','Efisiensi terbiasa','POS awam (s)','Efisiensi awam']
for i,h in enumerate(hd,start=1): ws3.cell(4,i,h)
style_header(ws3,4,1,6)
SUM=[
 ('Sederhana (9 tx)',34.7,26.9,'22,5%',33.5,'3,3%'),
 ('Kompleks (19 tx)',70.1,46.8,'33,2%',63.9,'8,9%'),
 ('SEMUA (28 tx)',58.7,40.4,'31,2%',54.1,'7,8%'),
 ('— Skenario kapabilitas —','','','','',''),
 ('Split-tender (1 meja 2 metode)',93.0,47.8,'48,6%',66.3,'28,7%'),
 ('Merge 2 meja (1 metode)',76.0,60.8,'19,9%',84.8,'−11,6%'),
 ('Merge 2 meja + split',110.0,66.0,'40,0%',93.5,'15,0%'),
]
rr=5
for s in SUM:
    nm,man,pm,em,pn,en=s
    ws3.cell(rr,1,nm)
    if man!='':
        ws3.cell(rr,2,fnum(man)).alignment=CTR; ws3.cell(rr,3,fnum(pm)).alignment=CTR
        ce=ws3.cell(rr,4,em); ce.alignment=CTR; ce.font=WIN
        ws3.cell(rr,5,fnum(pn)).alignment=CTR
        cn=ws3.cell(rr,6,en); cn.alignment=CTR; cn.font = LOSE if en.startswith('−') else WIN
        if nm.startswith('SEMUA'):
            for c in range(1,7): ws3.cell(rr,c).font=Font(bold=True)
    else:
        ws3.cell(rr,1).font=SUB
    rr+=1
box(ws3,4,1,rr-1,6)
ws3.cell(rr+1,1,'Catatan: merge sederhana saat awam JUSTRU lebih lambat (−11,6%) — disajikan apa adanya (netral).').font=NOTE

# =========================================================================
# SHEET 4 — Parameter & Asal Angka
# =========================================================================
ws4 = wb.create_sheet('Parameter & Asal Angka')
for col,w in zip('ABC',[46,16,52]): ws4.column_dimensions[col].width=w
ws4.cell(1,1,'Parameter Model & Asal Angka').font=TITLE
ws4.cell(2,1,'SEBELUM & SESUDAH dihitung = penjumlahan waktu per-langkah × komposisi nyata. Hanya latensi diukur; komposisi nyata dari buku.').font=NOTE
ws4.cell(2,1).alignment=WRAP; ws4.merge_cells('A2:C2'); ws4.row_dimensions[2].height=30
for i,h in enumerate(['Parameter','Nilai','Basis'],start=1): ws4.cell(4,i,h)
style_header(ws4,4,1,3)
PAR=[
 ('Tulis tangan / baris (manual)','5 s','tulis tangan ~13–25 WPM'),
 ('Tap menu POS / baris','6 s','hukum Fitts + scan (POS sedikit lebih lambat)'),
 ('Paket: tulis manual / modal POS','3 / 6 s','catat pilihan vs modal sub-pilihan'),
 ('Antar kertas waiter→kasir (manual)','2 s','POS: input langsung = 0'),
 ('Meja: catat manual / pilih POS','2 / 4 s','dine-in'),
 ('Hitung harga manual / baris','7 s','aritmetika manual ritel 3–7 s/baris (POS = 0)'),
 ('Finalisasi total (manual, >1 baris)','4 s','jumlah + tulis'),
 ('Merge: jumlah antar-meja / tandai','6 / 4 s','kalkulator + catat gabung (manual)'),
 ('Bayar tunai manual / POS','20 / 10 s','manual hitung+kembalian; POS auto-kembalian'),
 ('Bayar qris·edc·transfer (manual=POS)','12·16·14 s','non-tunai setara (tanpa payment gateway)'),
 ('Split: kalkulator sisa / catat (manual)','10 / 4 s','per slice'),
 ('Split POS: toggle / input slice','2 / 5 s','sisa otomatis'),
 ('Submit POS / struk PDF','2 / 0 (dikecualikan)','PDF = nilai-tambah, tak dihitung'),
 ('NOVICE_FACTOR (POS-awam, UI ×)','1,5','masa adaptasi; manual tanpa faktor'),
 ('Latensi tulis (DIUKUR)','0,14 s','+ mobile 4G 0,07 s (round-trip ke server prod)'),
]
rr=5
for p in PAR:
    ws4.cell(rr,1,p[0]).alignment=WRAP; ws4.cell(rr,2,p[1]).alignment=CTR; ws4.cell(rr,3,p[2]).alignment=WRAP
    rr+=1
box(ws4,4,1,rr-1,3)
rr+=1
ws4.cell(rr,1,'Sitasi: ketik mobile ~36 WPM (Palin dkk. 2019, Aalto); tulis tangan 13–25 WPM (SASC 2020); hukum Fitts (akuisisi target sentuh).').font=NOTE
ws4.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=3); ws4.row_dimensions[rr].height=28; ws4.cell(rr,1).alignment=WRAP
rr+=2
ws4.cell(rr,1,'KETERBATASAN (jujur):').font=SUB; rr+=1
for lim in [
 '1. Durasi = MODEL; akan dikalibrasi sampel stopwatch riil (manual & POS) oleh pegawai.',
 '2. Manual diasumsikan mahir (incumbent); POS dihitung 2 tingkat (terbiasa & awam).',
 '3. Latensi diukur dari klien ke server produksi; akses seluler dimodelkan (4G/3G).',
 '4. Cakupan = input → lunas (tanpa antrean/pikir pelanggan).',
 '5. Fitur rekomendasi kembalian tunai masih diasumsikan ada; split/merge = skenario kapabilitas.',
 '6. Sampel 28 tx (1 minggu) < 30 — keterbatasan generalisasi.',
]:
    ws4.cell(rr,1,lim).alignment=WRAP; ws4.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=3); rr+=1

wb.save(OUT)
print('SAVED', OUT)
